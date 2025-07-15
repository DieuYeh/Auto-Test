import tkinter as tk
from tkinter import filedialog, ttk
import re
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import subprocess

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PY 檔案與測試報告工具")
        self.geometry("900x700")

        self.py_files = {} 
        self.selected_cases_by_file = {} 
        self.tree_file_nodes = {} 
        self.last_py_folder = None 
        self.last_html_report_folder = None # 新增：用於儲存最近一次選擇的 HTML 報告資料夾
        self.last_excel_save_path = None # 新增：用於儲存最近一次保存 Excel 結果的路徑
        
        # 狀態訊息顯示區塊 (簡化為一行)
        self.status_label = tk.Label(self, text="準備就緒", bd=1, relief=tk.SUNKEN, anchor=tk.W, font=("Arial", 10))
        self.status_label.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(expand=True, fill="both", padx=10, pady=10)

        self.py_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.py_tab, text="PY 檔案處理")
        self.create_py_tab_widgets(self.py_tab)

        self.excel_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.excel_tab, text="Excel 報告處理")
        self.create_excel_tab_widgets(self.excel_tab)

    # 用於顯示狀態訊息的輔助函式 (簡化)
    def show_status_message(self, message, message_type="info"):
        color = "black" 
        if message_type == "warning":
            color = "orange"
        elif message_type == "error":
            color = "red"
        elif message_type == "success":
            color = "green"
        
        self.status_label.config(text=message, fg=color)
        # 5秒後自動清除訊息，恢復為"準備就緒"
        self.after(5000, lambda: self.status_label.config(text="準備就緒", fg="black"))

    def create_py_tab_widgets(self, parent_frame):
        # 檔案選擇區
        file_frame = tk.LabelFrame(parent_frame, text="PY 檔案載入", padx=10, pady=10)
        file_frame.pack(fill="x", padx=10, pady=5)

        self.file_list_label = tk.Label(file_frame, text="未載入任何 PY 檔案")
        self.file_list_label.pack(side=tk.LEFT, expand=True, fill="x")

        select_files_btn = tk.Button(file_frame, text="載入 PY 檔案", command=self.load_py_files)
        select_files_btn.pack(side=tk.RIGHT)

        # 搜尋設定區
        search_frame = tk.LabelFrame(parent_frame, text="測項分析", padx=10, pady=10)
        search_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(search_frame, text="備註: 將會分析PY內所含 \"test_case\"的測項，請符合名稱設計。", fg="blue").pack(side=tk.LEFT, padx=5)

        # 結果顯示區 (Treeview)
        result_frame = tk.LabelFrame(parent_frame, text="測項選擇結果", padx=10, pady=10)
        result_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.tree = ttk.Treeview(result_frame, columns=("No.", "Item Name", "Select"), show="headings")
        self.tree.heading("No.", text="No.", anchor=tk.CENTER)
        self.tree.heading("Item Name", text="項目名稱", anchor=tk.W)
        self.tree.heading("Select", text="選擇", anchor=tk.CENTER)

        self.tree.column("No.", width=50, anchor=tk.CENTER)
        self.tree.column("Item Name", width=400, anchor=tk.W)
        self.tree.column("Select", width=70, anchor=tk.CENTER)

        self.tree.tag_configure("file_node", background="#D3D3D3", foreground="blue", font=("", 9, "bold"))

        self.tree.pack(side=tk.LEFT, fill="both", expand=True)
        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
        self.tree.config(yscrollcommand=scrollbar.set)

        self.tree.bind("<ButtonRelease-1>", self.on_tree_click) 

        # 選項按鈕區
        option_frame = tk.Frame(parent_frame, padx=10, pady=5)
        option_frame.pack(fill="x", padx=10, pady=5)

        select_all_btn = tk.Button(option_frame, text="全選所有測項", command=self.select_all_test_items)
        select_all_btn.pack(side=tk.LEFT, padx=5)

        deselect_all_btn = tk.Button(option_frame, text="全取消所有測項", command=self.deselect_all_test_items)
        deselect_all_btn.pack(side=tk.LEFT, padx=5)

        self.selected_count_label = tk.Label(option_frame, text="已勾選測項: 0")
        self.selected_count_label.pack(side=tk.LEFT, padx=10)
        
        # 開啟資料夾按鈕 (PY分頁)
        self.open_py_folder_btn = tk.Button(option_frame, text="開啟PY資料夾", command=self.open_last_py_folder, state=tk.DISABLED)
        self.open_py_folder_btn.pack(side=tk.RIGHT, padx=5)

        # 匯出 Unit Plan 按鈕
        export_btn = tk.Button(option_frame, text="匯出 Unittest Plan", command=self.export_unittest_plan)
        export_btn.pack(side=tk.RIGHT)
        
    def open_last_py_folder(self):
        if self.last_py_folder and os.path.isdir(self.last_py_folder):
            try:
                if os.sys.platform == "win32":
                    os.startfile(self.last_py_folder)
                elif os.sys.platform == "darwin": 
                    subprocess.Popen(["open", self.last_py_folder])
                else: 
                    subprocess.Popen(["xdg-open", self.last_py_folder])
                self.show_status_message(f"已開啟PY資料夾: {self.last_py_folder}", "info")
            except Exception as e:
                self.show_status_message(f"無法開啟PY資料夾: {e}", "error")
        else:
            self.show_status_message("沒有可開啟的PY資料夾路徑。請先載入 PY 檔案。", "warning")

    def create_excel_tab_widgets(self, parent_frame):
        excel_settings_frame = tk.LabelFrame(parent_frame, text="Excel 設定", padx=10, pady=10)
        excel_settings_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(excel_settings_frame, text="讀取行數 (列A, 行B):").grid(row=0, column=0, sticky="w", pady=2)
        self.read_col_entry = tk.Entry(excel_settings_frame, width=5)
        self.read_col_entry.grid(row=0, column=1, padx=2, pady=2)
        self.read_row_entry = tk.Entry(excel_settings_frame, width=5)
        self.read_row_entry.grid(row=0, column=2, padx=2, pady=2)
        self.read_col_entry.insert(0, "A") 
        self.read_row_entry.insert(0, "1") 

        tk.Label(excel_settings_frame, text="寫入行數 (列A, 行B):").grid(row=1, column=0, sticky="w", pady=2)
        self.write_col_entry = tk.Entry(excel_settings_frame, width=5)
        self.write_col_entry.grid(row=1, column=1, padx=2, pady=2)
        self.write_row_entry = tk.Entry(excel_settings_frame, width=5)
        self.write_row_entry.grid(row=1, column=2, padx=2, pady=2)
        self.write_col_entry.insert(0, "B") 
        self.write_row_entry.insert(0, "1") 

        excel_buttons_frame = tk.Frame(parent_frame, padx=10, pady=5)
        excel_buttons_frame.pack(fill="x", padx=10, pady=5)

        load_testplan_btn = tk.Button(excel_buttons_frame, text="載入 Testplan (Excel)", command=self.load_testplan)
        load_testplan_btn.pack(side=tk.LEFT, padx=5)

        write_results_btn = tk.Button(excel_buttons_frame, text="寫入結果 (HTML -> Excel)", command=self.write_results_to_excel)
        write_results_btn.pack(side=tk.LEFT, padx=5)

        # 新增：開啟報告資料夾按鈕 (Excel分頁)
        self.open_report_folder_btn = tk.Button(excel_buttons_frame, text="開啟報告資料夾", command=self.open_last_report_folder, state=tk.DISABLED)
        self.open_report_folder_btn.pack(side=tk.RIGHT, padx=5)

    def open_last_report_folder(self):
        target_folder = None
        if self.last_excel_save_path and os.path.isdir(os.path.dirname(self.last_excel_save_path)):
            target_folder = os.path.dirname(self.last_excel_save_path)
            message_type = "Excel結果"
        elif self.last_html_report_folder and os.path.isdir(self.last_html_report_folder):
            target_folder = self.last_html_report_folder
            message_type = "HTML報告"
        else:
            self.show_status_message("沒有可開啟的報告資料夾路徑。請先載入 HTML 報告或保存 Excel 結果。", "warning")
            return

        try:
            if os.sys.platform == "win32":
                os.startfile(target_folder)
            elif os.sys.platform == "darwin": 
                subprocess.Popen(["open", target_folder])
            else: 
                subprocess.Popen(["xdg-open", target_folder])
            self.show_status_message(f"已開啟{message_type}資料夾: {target_folder}", "info")
        except Exception as e:
            self.show_status_message(f"無法開啟報告資料夾: {e}", "error")

    def load_py_files(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("Python files", "*.py")])
        if file_paths:
            self.py_files.clear() 
            self.last_py_folder = os.path.dirname(file_paths[0]) 
            
            for file_path in file_paths:
                module_name = os.path.splitext(os.path.basename(file_path))[0]
                self.py_files[file_path] = {'module_name': module_name, 'test_class_name': None}
            
            if len(self.py_files) > 0:
                displayed_names = ", ".join([os.path.basename(path) for path in self.py_files.keys()])
                self.file_list_label.config(text=f"已選擇: {displayed_names}")
                self.analyze_all_py_files() 
                self.open_py_folder_btn.config(state=tk.NORMAL) # 更新按鈕名稱
                self.show_status_message(f"已載入 {len(file_paths)} 個 PY 檔案。", "success")
            else:
                self.file_list_label.config(text="未載入任何 PY 檔案")
                self.open_py_folder_btn.config(state=tk.DISABLED) # 更新按鈕名稱
                self.show_status_message("未載入任何 PY 檔案。", "warning")
        else:
            self.show_status_message("取消載入 PY 檔案。", "info")

    def analyze_all_py_files(self):
        if not self.py_files:
            self.show_status_message("請先載入 PY 檔案！", "warning")
            return

        self.selected_cases_by_file.clear()
        self.tree.delete(*self.tree.get_children()) 
        self.tree_file_nodes.clear() 

        total_cases_found = 0
        overall_no = 1 

        test_case_pattern = re.compile(r'def\s+(test_case[a-zA-Z0-9_]+)\s*\(self\):') 
        test_class_pattern = re.compile(r'class\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*\(\s*unittest\.TestCase\s*\):')

        for py_file_path, file_info in self.py_files.items():
            file_display_name = os.path.basename(py_file_path) 
            file_node_id = py_file_path 

            self.tree.insert("", "end", iid=file_node_id, 
                             values=("", file_display_name, "☐"), 
                             tags=("file_node", "checkbox")) 
            self.tree_file_nodes[py_file_path] = file_node_id
            
            self.tree.item(file_node_id, open=True) 
            
            self.selected_cases_by_file[py_file_path] = {}

            try:
                with open(py_file_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                class_matches = list(test_class_pattern.finditer(content))
                if class_matches:
                    detected_class_name = class_matches[0].group(1)
                    file_info['test_class_name'] = detected_class_name
                    self.tree.item(file_node_id, values=("", f"{file_display_name} (Class: {detected_class_name})", "☐"))
                else:
                    file_info['test_class_name'] = "MyTestCase" 
                    self.tree.item(file_node_id, values=("", f"{file_display_name} (未偵測到測試類別, 預設 MyTestCase)", "☐"))
                    
                matches = test_case_pattern.finditer(content)
                
                file_cases_count = 0
                for match in matches:
                    case_name = match.group(1) 
                    if case_name not in self.selected_cases_by_file[py_file_path]:
                        var = tk.BooleanVar(value=False) 
                        self.selected_cases_by_file[py_file_path][case_name] = var
                        self.tree.insert(file_node_id, "end", iid=f"{file_node_id}-{case_name}", 
                                         values=(overall_no, case_name, "☐"), tags=("checkbox",))
                        total_cases_found += 1
                        file_cases_count += 1
                        overall_no += 1 
                
                if file_cases_count == 0:
                    current_values = list(self.tree.item(file_node_id, "values"))
                    if "(Class:" in current_values[1]:
                        current_values[1] = current_values[1].replace(")", ", 無測項)") 
                    else:
                        current_values[1] = f"{os.path.basename(py_file_path)} (無測項)" 
                    self.tree.item(file_node_id, values=current_values)

            except Exception as e:
                self.tree.item(file_node_id, values=("", f"{file_display_name} (讀取失敗: {e})", "☐"))
                self.show_status_message(f"讀取或搜尋檔案 '{file_display_name}' 時發生錯誤: {e}", "error")

        self.update_selected_count_label()
        if total_cases_found > 0:
            self.show_status_message(f"已成功分析所有 PY 檔案，共找到 {total_cases_found} 個測試案例。", "success")
        else:
            self.show_status_message("未在載入的 PY 檔案中找到任何測試案例。", "warning")

    def on_tree_click(self, event):
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        column = self.tree.identify_column(event.x)
        if column == "#3": 
            tags = self.tree.item(item_id, "tags")
            
            if "file_node" in tags:
                py_file_path = item_id 
                
                all_children = self.tree.get_children(item_id)
                if not all_children: 
                    return

                any_unselected = False
                for child_id in all_children:
                    parts = child_id.split('-', 1) 
                    if len(parts) < 2: continue 
                    
                    child_py_path = parts[0]
                    child_case_name = parts[1]
                    
                    if child_py_path == py_file_path: 
                        if child_case_name in self.selected_cases_by_file[child_py_path]:
                            if not self.selected_cases_by_file[child_py_path][child_case_name].get():
                                any_unselected = True
                                break
                
                new_state = any_unselected 

                for child_id in all_children:
                    parts = child_id.split('-', 1)
                    if len(parts) < 2: continue
                    child_py_path = parts[0]
                    child_case_name = parts[1]

                    if child_py_path == py_file_path:
                        self.selected_cases_by_file[child_py_path][case_name].set(new_state)
                        self.update_checkbox_display(child_id, new_state)
                
                self.update_file_node_checkbox_display(py_file_path)
                self.update_selected_count_label()

            else:
                parent_id = self.tree.parent(item_id)
                if not parent_id: return 

                py_file_path = parent_id
                case_name = item_id.split('-', 1)[1] 

                if py_file_path in self.selected_cases_by_file and \
                   case_name in self.selected_cases_by_file[py_file_path]:
                    current_var = self.selected_cases_by_file[py_file_path][case_name]
                    current_var.set(not current_var.get()) 
                    self.update_checkbox_display(item_id, current_var.get())
                    self.update_file_node_checkbox_display(py_file_path) 
                    self.update_selected_count_label()

    def update_checkbox_display(self, item_id, is_selected):
        current_values = list(self.tree.item(item_id, "values"))
        current_values[2] = "☑" if is_selected else "☐"
        self.tree.item(item_id, values=current_values)

    def update_file_node_checkbox_display(self, py_file_path):
        if py_file_path not in self.tree_file_nodes:
            return

        file_node_id = self.tree_file_nodes[py_file_path]
        file_cases = self.selected_cases_by_file.get(py_file_path, {})
        
        if not file_cases: 
            current_values = list(self.tree.item(file_node_id, "values"))
            current_values[2] = "☐" 
            self.tree.item(file_node_id, values=current_values)
            return

        total_cases = len(file_cases)
        selected_cases = sum(1 for var in file_cases.values() if var.get())

        current_values = list(self.tree.item(file_node_id, "values"))
        if selected_cases == total_cases:
            current_values[2] = "☑" 
        elif selected_cases > 0:
            current_values[2] = "■" 
        else:
            current_values[2] = "☐" 

        self.tree.item(file_node_id, values=current_values)


    def update_selected_count_label(self):
        count = 0
        for file_cases in self.selected_cases_by_file.values():
            count += sum(1 for var in file_cases.values() if var.get())
        self.selected_count_label.config(text=f"已勾選測項: {count}") 

    def select_all_test_items(self):
        for py_file_path, file_cases in self.selected_cases_by_file.items():
            for case_name, var in file_cases.items():
                var.set(True)
                item_id = f"{py_file_path}-{case_name}"
                self.update_checkbox_display(item_id, True)
            self.update_file_node_checkbox_display(py_file_path) 
        self.update_selected_count_label()
        self.show_status_message("已全選所有測試案例。", "info")

    def deselect_all_test_items(self):
        for py_file_path, file_cases in self.selected_cases_by_file.items():
            for case_name, var in file_cases.items():
                var.set(False)
                item_id = f"{py_file_path}-{case_name}"
                self.update_checkbox_display(item_id, False)
            self.update_file_node_checkbox_display(py_file_path) 
        self.update_selected_count_label()
        self.show_status_message("已取消全選所有測試案例。", "info")

    def export_unittest_plan(self):
        selected_cases_by_module = {}

        for py_file_path, file_cases in self.selected_cases_by_file.items():
            file_info = self.py_files[py_file_path]
            module_name = file_info['module_name']
            test_class_name = file_info['test_class_name'] if file_info['test_class_name'] else "MyTestCase"

            for case_name, var in file_cases.items():
                if var.get():
                    if module_name not in selected_cases_by_module:
                        selected_cases_by_module[module_name] = {'class_name': test_class_name, 'cases': []}
                    selected_cases_by_module[module_name]['cases'].append(case_name)

        if not selected_cases_by_module:
            self.show_status_message("請至少選擇一個測試案例！", "warning")
            return

        initial_dir = self.last_py_folder if self.last_py_folder and os.path.isdir(self.last_py_folder) else None
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".py",
            filetypes=[("Python files", "*.py"), ("All files", "*.*")],
            initialfile="Unittest_plan.py",
            initialdir=initial_dir
        )

        if not file_path:
            self.show_status_message("已取消匯出 Unittest Plan。", "info")
            return

        output_content = [
            "import unittest\n",
            "import HTMLTestRunner # type: ignore\n",
            "import os\n",
            "\n"
        ]

        for mod_name, info in sorted(selected_cases_by_module.items()):
            output_content.append(f"from {mod_name} import {info['class_name']}\n")

        output_content.append("\n")
        output_content.append("if __name__ == '__main__':\n")
        output_content.append("    # 確保測試報告目錄存在\n")
        output_content.append("    report_dir = 'D:/SeleniumProject/test_reports'\n")
        output_content.append("    os.makedirs(report_dir, exist_ok=True)\n")
        output_content.append("\n")

        for mod_name, info in sorted(selected_cases_by_module.items()):
            class_name = info['class_name']
            output_content.append(f"    print(f\"\\n--- Running tests for {mod_name}.py ---\")\n")
            output_content.append(f"    suite_{mod_name} = unittest.TestSuite()\n")
            for case_name in info['cases']:
                output_content.append(f"    suite_{mod_name}.addTest({class_name}('{case_name}'))\n")
            
            output_content.append(f"    report_file_{mod_name} = os.path.join(report_dir, f'Report_{mod_name}.html')\n")
            output_content.append(f"    with open(report_file_{mod_name}, 'wb') as f:\n")
            output_content.append(f"        runner_{mod_name} = HTMLTestRunner.HTMLTestRunner(\n")
            output_content.append("            stream=f,\n")
            output_content.append(f"            title='{mod_name} Test Report',\n")
            output_content.append("            description='Individual test report for {mod_name}.py'\n")
            output_content.append("        )\n")
            output_content.append(f"        runner_{mod_name}.run(suite_{mod_name})\n")
            output_content.append(f"    print(f\"Test report for {mod_name}.py saved to {{report_file_{mod_name}}}\")\n")
            output_content.append("\n")
        
        output_content.append("    print(\"\\n--- All selected test suites have been executed and reported. ---\")\n")

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.writelines(output_content)
            self.show_status_message(f"Unittest Plan 已成功匯出到 '{os.path.basename(file_path)}'，報告將生成在 'D:/SeleniumProject/test_reports'。可點選開啟PY資料夾。", "success")
        except Exception as e:
            self.show_status_message(f"匯出檔案時發生錯誤: {e}", "error")

    def load_testplan(self):
        excel_file_path = filedialog.askopenfilenames(
            filetypes=[
                ("Excel files (XLSX)", "*.xlsx"), 
                ("Excel files (XLS)", "*.xls")
            ]
        )
        if excel_file_path:
            result_dir = "Result"
            os.makedirs(result_dir, exist_ok=True) 

            for f_path in excel_file_path: 
                dest_path = os.path.join(result_dir, os.path.basename(f_path))
                try:
                    shutil.copy(f_path, dest_path)
                    self.show_status_message(f"已成功載入 Testplan: {os.path.basename(f_path)}", "success")
                except Exception as e:
                    self.show_status_message(f"複製檔案 '{os.path.basename(f_path)}' 時發生錯誤: {e}", "error")
            if not excel_file_path: 
                self.show_status_message("未選擇任何 Excel Testplan 檔案。", "warning") 
        else:
            self.show_status_message("取消載入 Excel Testplan。", "info")


    def write_results_to_excel(self):
        html_dir = filedialog.askdirectory(title="選擇包含 HTML 報告的資料夾")
        if not html_dir:
            self.show_status_message("取消選擇 HTML 報告資料夾。", "info")
            return
        self.last_html_report_folder = html_dir # 儲存選擇的 HTML 資料夾路徑
        self.open_report_folder_btn.config(state=tk.NORMAL) # 啟用按鈕

        excel_files_in_result = [f for f in os.listdir("Result") if f.endswith((".xlsx", ".xls"))]
        if not excel_files_in_result:
            self.show_status_message("Result 資料夾中沒有找到 Excel Testplan 檔案！請先載入。", "warning")
            return

        testplan_path_in_result = os.path.join("Result", excel_files_in_result[0])

        try:
            workbook = load_workbook(testplan_path_in_result)
            sheet = workbook.active 

            read_col = self.read_col_entry.get().upper()
            read_row = int(self.read_row_entry.get()) - 1 
            write_col = self.write_col_entry.get().upper()
            write_row = int(self.write_row_entry.get()) - 1

            if not read_col.isalpha() or not write_col.isalpha():
                self.show_status_message("讀取/寫入欄位必須是字母 (例如 A, B)！", "error")
                return
            if read_row < 0 or write_row < 0:
                self.show_status_message("讀取/寫入行數必須是正整數！", "error")
                return

            read_col_idx = ord(read_col) - ord('A') 
            write_col_idx = ord(write_col) - ord('A')

            html_files = [f for f in os.listdir(html_dir) if f.endswith(".html")]
            if not html_files:
                self.show_status_message("選擇的資料夾中沒有找到 HTML 報告檔案。", "warning")
                return

            all_html_results = {}
            for html_file in html_files:
                file_path = os.path.join(html_dir, html_file)
                
                content = None
                encodings_to_try = ['utf-8', 'gbk', 'cp950', 'latin-1'] 
                for encoding in encodings_to_try:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            content = f.read()
                        break 
                    except UnicodeDecodeError:
                        pass 
                    except Exception as e:
                        break 

                if content is None:
                    self.show_status_message(f"無法成功讀取 HTML 檔案 '{html_file}'，請檢查其編碼或檔案完整性。", "warning")
                    continue 
                
                try:
                    soup = BeautifulSoup(content, 'html.parser')
                    results = self.parse_html_report(soup)
                    for item in results:
                        all_html_results[item['name']] = item['result'] 
                except Exception as e:
                    self.show_status_message(f"解析 HTML 檔案 '{html_file}' 內容時發生錯誤: {e}", "warning")

            results_written_count = 0
            for row_idx, row in enumerate(sheet.iter_rows()):
                if row_idx < read_row: 
                    continue

                testcase_name_cell = sheet.cell(row=row_idx + 1, column=read_col_idx + 1) 
                testcase_name = str(testcase_name_cell.value).strip() if testcase_name_cell.value else ""

                if testcase_name: 
                    if testcase_name in all_html_results:
                        result_to_write = all_html_results[testcase_name]
                        sheet.cell(row=row_idx + 1, column=write_col_idx + 1, value=result_to_write)
                        results_written_count += 1

            original_filename = os.path.basename(testplan_path_in_result)
            name, ext = os.path.splitext(original_filename)
            default_save_filename = f"{name}_results{ext}"

            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_save_filename
            )

            if not save_path: 
                self.show_status_message("已處理的 Excel 檔案未保存。", "warning")
                return

            workbook.save(save_path)
            self.last_excel_save_path = save_path # 儲存保存的 Excel 路徑
            self.show_status_message(f"測試結果已成功寫入並保存到:\n{save_path}", "success")
            self.open_report_folder_btn.config(state=tk.NORMAL) # 啟用按鈕

        except Exception as e:
            self.show_status_message(f"寫入結果到 Excel 時發生錯誤: {e}", "error")

    def parse_html_report(self, soup):
        all_test_results = []
        for tr_tag in soup.find_all('tr', class_=['hiddenRow', 'none']):
            test_info = {}

            name_tag_container = tr_tag.find('td', class_=['passCase', 'failCase'])
            if name_tag_container:
                name_tag = name_tag_container.find('div', class_='testcase').find('a', class_='popup_link')
                if name_tag:
                    test_info['name'] = name_tag.get_text(strip=True)
                else:
                    test_info['name'] = "N/A (名稱連結未找到)"
            else:
                test_info['name'] = "N/A (名稱欄位未找到)"

            result_tag_container = tr_tag.find('td', align='center')
            if result_tag_container:
                result_link = result_tag_container.find('button').find('a', class_='popup_link')
                if result_link:
                    test_info['result'] = result_link.get_text(strip=True)
                else:
                    test_info['result'] = "N/A (結果連結未找到)"
            else:
                test_info['result'] = "N/A (結果欄位未找到)"

            if test_info['name'].startswith("test_case") and test_info['result'] not in ["N/A (結果連結未找到)", "N/A (結果欄位未找到)"]:
                all_test_results.append(test_info)
        return all_test_results


if __name__ == "__main__":
    app = App()
    app.mainloop()